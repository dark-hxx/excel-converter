"""Excel 转换器主入口：苹果系统亮色风格 UI。

UI 框架：customtkinter
通用模板映射业务逻辑：保留原实现
所有 messagebox / simpledialog 已替换为 Banner / CTkInputDialog
"""
from datetime import datetime
import json
import os
import sys
from tkinter import filedialog, ttk

import customtkinter as ctk
from openpyxl import load_workbook
from openpyxl.styles import numbers
import pandas as pd

from apple_theme import (
    apply_apple_theme,
    font_ui, font_title, font_mono,
    BUTTON_PRIMARY, BUTTON_SECONDARY, BUTTON_PLAIN,
    CARD_STYLE, ENTRY_STYLE, TEXTBOX_STYLE,
    BLUE, RED, GREEN, ORANGE,
    TEXT_PRIMARY, TEXT_SECONDARY, HOVER_BG,
    WINDOW_BG, CARD_BG,
    show_banner, transparent_frame,
)
from bank_converter import open_bank_converter_window, open_batch_converter_window
from utils import build_unique_save_path, center_window, sanitize_filename_part


# ---------------- 路径常量 ----------------

if getattr(sys, 'frozen', False):
    base_dir = os.path.dirname(sys.executable)
else:
    base_dir = os.getcwd()

HISTORY_TEMPLATES_FILE = os.path.join(base_dir, 'history_templates.json')
HISTORY_MAPPINGS_FILE = os.path.join(base_dir, 'history_mappings.json')


def save_text_excel(dataframe, save_path):
    output_df = dataframe.copy()
    for col in output_df.columns:
        output_df[col] = output_df[col].astype(str).replace('nan', '')

    output_df.to_excel(save_path, index=False, sheet_name='Sheet1', engine='openpyxl')
    wb = load_workbook(save_path)
    try:
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.number_format = numbers.FORMAT_TEXT
        wb.save(save_path)
    finally:
        wb.close()


def build_adjusted_split_info(split_info, template_to_file_mapping,
                              template_columns_ordered):
    reverse_column_mapping = {
        v.strip().lower(): k.strip().lower()
        for k, v in template_to_file_mapping.items()
    }
    adjusted_split_info = {}
    for file_col, split_symbol in split_info.items():
        tpl_col_lower = reverse_column_mapping.get(file_col.strip().lower())
        if tpl_col_lower:
            for orig_col in template_columns_ordered:
                if orig_col.strip().lower() == tpl_col_lower:
                    adjusted_split_info[orig_col] = split_symbol
                    break
        else:
            log(f"警告: split_info 中的列 '{file_col}' 未在映射中找到", 'warning')
    return adjusted_split_info


def apply_split_info(dataframe, adjusted_split_info):
    if not adjusted_split_info:
        return dataframe

    new_rows = []
    for index, row in dataframe.iterrows():
        try:
            split_values_dict = {}
            max_length = 1
            for col_name, sym in adjusted_split_info.items():
                if col_name in row.index:
                    if pd.notna(row[col_name]):
                        values = str(row[col_name]).split(sym)
                        values = [v.strip().strip(sym).strip() for v in values]
                        values = [v for v in values if v]
                        if values:
                            split_values_dict[col_name] = values
                            max_length = max(max_length, len(values))
                        else:
                            split_values_dict[col_name] = [str(row[col_name])]
                    else:
                        split_values_dict[col_name] = ['']
            if split_values_dict:
                for i in range(max_length):
                    new_row = row.copy()
                    for col_name, values in split_values_dict.items():
                        new_row[col_name] = values[i] if i < len(values) else ''
                    new_rows.append(new_row)
                log(f"信息: 行 {index} 拆分为 {max_length} 行")
            else:
                new_rows.append(row)
        except Exception as row_error:
            log(f"错误: 处理行 {index} 时出错: {row_error}", 'error')
            new_rows.append(row)
    return pd.DataFrame(new_rows)


def convert_date_format(fmt):
    if not fmt:
        return fmt
    replacements = [
        ('yyyy', '%Y'), ('yy', '%y'), ('MM', '%m'),
        ('dd', '%d'), ('HH', '%H'), ('mm', '%M'), ('ss', '%S'),
    ]
    result = fmt
    for old, new in replacements:
        result = result.replace(old, new)
    return result


def apply_date_formats(dataframe, date_format_info):
    for col, fmt_config in date_format_info.items():
        if col in dataframe.columns:
            input_fmt = convert_date_format(fmt_config.get('input', ''))
            output_fmt = convert_date_format(fmt_config.get('output', ''))
            if input_fmt and output_fmt:
                def convert_date(val, in_fmt=input_fmt, out_fmt=output_fmt):
                    if pd.isna(val) or str(val).strip() == '':
                        return ''
                    try:
                        dt = datetime.strptime(str(val).strip(), in_fmt)
                        return dt.strftime(out_fmt)
                    except ValueError:
                        return str(val)
                dataframe[col] = dataframe[col].apply(convert_date)
    return dataframe


# ---------------- 全局状态（通用模板映射用） ----------------

save_dir = None
template_df = None
column_mapping = None

# UI 引用（main 块中初始化）
root = None
log_text = None
banner_area = None


# ---------------- 日志工具 ----------------

def _detect_level(msg):
    if '错误' in msg or '失败' in msg or '异常' in msg:
        return 'error'
    if '警告' in msg or '跳过' in msg:
        return 'warning'
    if '成功' in msg or '完成' in msg:
        return 'success'
    return 'info'


def log(msg, level=None):
    """写日志，按 level 染色。level=None 时按内容自动检测。"""
    if log_text is None:
        return
    if level is None:
        level = _detect_level(msg)
    line = msg if msg.endswith('\n') else msg + '\n'
    if level == 'info':
        log_text.insert('end', line)
    else:
        log_text.insert('end', line, level)
    log_text.see('end')
    log_text.update_idletasks()


# ---------------- 历史记录 IO ----------------

def load_history_templates():
    try:
        with open(HISTORY_TEMPLATES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return []


def save_history_template(template_path):
    templates = load_history_templates()
    if template_path not in templates:
        templates.append(template_path)
    with open(HISTORY_TEMPLATES_FILE, 'w', encoding='utf-8') as f:
        json.dump(templates, f, ensure_ascii=False, indent=4)


def load_history_mappings():
    try:
        with open(HISTORY_MAPPINGS_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
            if isinstance(data, list):
                return {f'mapping_{i}': item for i, item in enumerate(data)}
            return data
    except FileNotFoundError:
        return {}


# ---------------- 模板/映射选择 ----------------

def select_template():
    global template_df
    template_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx;*.xls")])
    if not template_path:
        return
    try:
        template_df = pd.read_excel(template_path, sheet_name=0, dtype=str)
        log(f"模板文件已加载: {template_path}", 'success')
        save_history_template(template_path)
    except Exception as e:
        show_banner(banner_area, f"加载模板出错：{e}", 'error')
        log(f"加载模板出错: {e}", 'error')


def set_column_mapping():
    global column_mapping
    if template_df is None:
        show_banner(banner_area, "请先选择模板文件", 'warning')
        return

    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx;*.xls")])
    if not file_path:
        return

    try:
        df = pd.read_excel(file_path, sheet_name=0, dtype=str)
        file_columns = df.columns.tolist()
    except Exception as e:
        show_banner(banner_area, f"读取文件出错：{e}", 'error')
        log(f"读取文件 {file_path} 出错: {e}", 'error')
        return

    mapping_window = ctk.CTkToplevel(root)
    mapping_window.title("设置列映射")
    mapping_window.configure(fg_color=WINDOW_BG)
    mapping_window.transient(root)
    mapping_window.grab_set()

    ctk.CTkLabel(mapping_window, text='设置列映射', font=font_title(15),
                 text_color=TEXT_PRIMARY, anchor='w'
                 ).pack(fill='x', padx=20, pady=(16, 8))

    container = ctk.CTkScrollableFrame(mapping_window, **CARD_STYLE)
    container.pack(fill='both', expand=True, padx=16, pady=(0, 12))

    date_formats = [
        "", "yyyyMMdd", "yyyy/MM/dd", "yyyy-MM-dd", "yyyy年MM月dd日",
        "dd/MM/yyyy", "MM/dd/yyyy", "yyyy-MM-dd HH:mm:ss",
        "yyyy/MM/dd HH:mm:ss", "yyyyMMddHHmmss",
    ]

    template_columns = list(template_df.columns)
    headers = ['模板列', '文件列', '分隔符', '输入时间格式', '输出时间格式']
    for col_idx, header in enumerate(headers):
        ctk.CTkLabel(container, text=header, font=font_ui(12, 'bold'),
                     text_color=TEXT_SECONDARY
                     ).grid(row=0, column=col_idx, padx=8, pady=(6, 8), sticky='w')

    combo_boxes, entry_boxes, input_date_combos, output_date_combos = [], [], [], []
    for i, col in enumerate(template_columns):
        ctk.CTkLabel(container, text=col, font=font_ui(12),
                     text_color=TEXT_PRIMARY, anchor='w'
                     ).grid(row=i + 1, column=0, padx=8, pady=4, sticky='w')

        combo = ttk.Combobox(container, values=file_columns, width=20)
        combo.grid(row=i + 1, column=1, padx=8, pady=4, sticky='w')
        combo_boxes.append(combo)

        entry = ctk.CTkEntry(container, width=100, **ENTRY_STYLE)
        entry.grid(row=i + 1, column=2, padx=8, pady=4, sticky='w')
        entry_boxes.append(entry)

        in_combo = ttk.Combobox(container, values=date_formats, width=20)
        in_combo.grid(row=i + 1, column=3, padx=8, pady=4, sticky='w')
        input_date_combos.append(in_combo)

        out_combo = ttk.Combobox(container, values=date_formats, width=20)
        out_combo.grid(row=i + 1, column=4, padx=8, pady=4, sticky='w')
        output_date_combos.append(out_combo)

    def do_save():
        global column_mapping
        dialog = ctk.CTkInputDialog(text="请输入当前映射关系的名称：", title="保存映射")
        name = dialog.get_input()
        if not name:
            show_banner(banner_area, "未输入映射名称，未保存", 'warning')
            return
        column_mapping = {}
        split_info = {}
        date_format_info = {}
        for i, col in enumerate(template_columns):
            sel = combo_boxes[i].get()
            split = entry_boxes[i].get()
            in_fmt = input_date_combos[i].get()
            out_fmt = output_date_combos[i].get()
            if sel:
                column_mapping[col] = sel
                if split:
                    split_info[sel] = split
                if in_fmt and out_fmt:
                    date_format_info[col] = {"input": in_fmt, "output": out_fmt}
        column_mapping['split_info'] = split_info
        column_mapping['date_format_info'] = date_format_info
        mapping_window.destroy()
        log("列映射已成功保存", 'success')
        history = load_history_mappings()
        history[name] = column_mapping
        with open(HISTORY_MAPPINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=4)

    save_row = transparent_frame(mapping_window)
    save_row.pack(pady=(0, 16))
    ctk.CTkButton(save_row, text='保存映射', command=do_save,
                  width=180, font=font_ui(13, 'bold'), **BUTTON_PRIMARY
                  ).pack()

    # 居中
    mapping_window.update_idletasks()
    w, h = 780, min(620, 140 + 40 * len(template_columns) + 80)
    sx = mapping_window.winfo_screenwidth()
    sy = mapping_window.winfo_screenheight()
    mapping_window.geometry(f'{w}x{h}+{(sx-w)//2}+{(sy-h)//2}')


def select_save_directory():
    global save_dir
    chosen = filedialog.askdirectory()
    if chosen:
        save_dir = chosen
        log(f"保存目录已设置为: {save_dir}", 'success')


# ---------------- 主转换流程 ----------------

def convert_excel_files():
    if template_df is None or column_mapping is None:
        show_banner(banner_area, "请先选择模板并设置列映射", 'warning')
        return
    if save_dir is None:
        show_banner(banner_area, "请先选择保存目录", 'warning')
        return

    file_paths = filedialog.askopenfilenames(
        filetypes=[("Excel files", "*.xlsx;*.xls")])
    if not file_paths:
        return

    split_info = column_mapping.get('split_info', {})
    date_format_info = column_mapping.get('date_format_info', {})
    template_to_file_mapping = {
        k: v for k, v in column_mapping.items()
        if k not in ('split_info', 'date_format_info')
    }

    success_count = 0
    error_count = 0

    for file_path in file_paths:
        try:
            sheet_names = pd.ExcelFile(file_path).sheet_names

            for sheet_name in sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)

                log(f"调试: 文件 {file_path} 工作表 {sheet_name} - 开始处理")

                file_columns_ordered = [
                    template_to_file_mapping[k] for k in template_to_file_mapping.keys()
                ]
                missing_columns = [c for c in file_columns_ordered if c not in df.columns]
                if missing_columns:
                    log(f"错误: 文件 {file_path} 工作表 {sheet_name} 缺少列: {missing_columns}",
                        'error')
                    error_count += 1
                    continue

                mapped_df = df[file_columns_ordered].copy()
                template_columns_ordered = list(template_to_file_mapping.keys())
                mapped_df.columns = template_columns_ordered

                adjusted_split_info = build_adjusted_split_info(
                    split_info, template_to_file_mapping, template_columns_ordered
                )
                mapped_df = apply_split_info(mapped_df, adjusted_split_info)

                if date_format_info:
                    mapped_df = apply_date_formats(mapped_df, date_format_info)

                all_template_columns = list(template_df.columns)
                for col in all_template_columns:
                    if col not in mapped_df.columns:
                        mapped_df[col] = ''
                mapped_df = mapped_df[all_template_columns]

                file_name = sanitize_filename_part(
                    os.path.splitext(os.path.basename(file_path))[0]
                )
                if len(sheet_names) > 1:
                    file_name += f'_{sanitize_filename_part(sheet_name)}'
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                save_path = build_unique_save_path(save_dir, f'{file_name}_{timestamp}')

                try:
                    save_text_excel(mapped_df, save_path)
                    log(f"成功: 文件 {file_path} 工作表 {sheet_name} 转换完成 → {save_path}",
                        'success')
                    success_count += 1
                except PermissionError:
                    msg = f"无法保存 {save_path}：文件可能被 Excel 占用"
                    log(msg, 'error')
                    show_banner(banner_area, msg, 'error')
                    error_count += 1
                    continue
                except Exception as save_error:
                    log(f"错误: 保存 {save_path} 出错: {save_error}", 'error')
                    show_banner(banner_area, f"保存出错: {save_error}", 'error')
                    error_count += 1
                    continue

        except Exception as e:
            log(f"错误: 处理文件 {file_path} 时出错: {e}", 'error')
            show_banner(banner_area,
                        f"处理 {os.path.basename(file_path)} 出错: {e}", 'error')
            error_count += 1

    if error_count == 0:
        show_banner(banner_area, f"全部完成：成功转换 {success_count} 个工作表", 'success')
    else:
        show_banner(banner_area,
                    f"完成：成功 {success_count}，失败 {error_count}", 'warning')


# ---------------- 历史记录展示 ----------------

def show_history_templates():
    templates = load_history_templates()
    if not templates:
        show_banner(banner_area, "暂无历史模板记录", 'info')
        return
    _show_history_list(templates, "历史模板",
                       lambda p: use_history_template(p),
                       truncate=60)


def show_history_mappings():
    mappings = load_history_mappings()
    if not mappings:
        show_banner(banner_area, "暂无历史映射记录", 'info')
        return
    _show_history_list(list(mappings.keys()), "历史映射",
                       lambda name: use_history_mapping(mappings[name]),
                       truncate=60)


def _show_history_list(items, title, on_click, truncate=80):
    win = ctk.CTkToplevel(root)
    win.title(title)
    win.configure(fg_color=WINDOW_BG)
    win.transient(root)
    win.grab_set()

    w, h = 580, 440
    sx, sy = win.winfo_screenwidth(), win.winfo_screenheight()
    win.geometry(f'{w}x{h}+{(sx-w)//2}+{(sy-h)//2}')

    ctk.CTkLabel(win, text=title, font=font_title(15),
                 text_color=TEXT_PRIMARY, anchor='w'
                 ).pack(fill='x', padx=20, pady=(16, 8))

    list_frame = ctk.CTkScrollableFrame(win, **CARD_STYLE)
    list_frame.pack(fill='both', expand=True, padx=16, pady=(0, 16))

    for item in items:
        display = item if len(item) <= truncate else '…' + item[-(truncate - 1):]

        def make_handler(val=item):
            def handle_click():
                on_click(val)
                win.destroy()
            return handle_click

        ctk.CTkButton(
            list_frame, text=display, command=make_handler(),
            anchor='w', font=font_ui(12),
            fg_color='transparent', text_color=TEXT_PRIMARY,
            hover_color=HOVER_BG, corner_radius=6, height=36,
        ).pack(fill='x', padx=4, pady=2)


def use_history_template(template_path):
    global template_df
    try:
        template_df = pd.read_excel(template_path, sheet_name=0, dtype=str)
        log(f"历史模板已加载: {template_path}", 'success')
        save_history_template(template_path)
    except Exception as e:
        show_banner(banner_area, f"加载历史模板出错: {e}", 'error')
        log(f"加载历史模板出错: {e}", 'error')


def use_history_mapping(mapping):
    global column_mapping
    column_mapping = mapping
    log("历史映射已应用", 'success')


# ---------------- 主窗口构建 ----------------

def build_main_window():
    global root, log_text, banner_area

    root = ctk.CTk()
    apply_apple_theme(root)
    root.title('Excel 转换器')
    center_window(root, 580, 820)

    # 顶部 banner 区（show_banner 注入位置）
    banner_area = transparent_frame(root)
    banner_area.pack(fill='x', side='top')

    # 标题栏
    title_bar = transparent_frame(root, height=44)
    title_bar.pack(fill='x', padx=20, pady=(16, 8))
    ctk.CTkLabel(title_bar, text='Excel 转换器',
                 font=font_title(18), text_color=TEXT_PRIMARY
                 ).pack(side='left')

    # 银行流水转换卡片
    bank_card = ctk.CTkFrame(root, **CARD_STYLE)
    bank_card.pack(fill='x', padx=20, pady=(0, 12))

    ctk.CTkLabel(bank_card, text='银行流水转换',
                 font=font_title(14), text_color=TEXT_PRIMARY, anchor='w'
                 ).pack(fill='x', padx=16, pady=(14, 2))
    ctk.CTkLabel(bank_card,
                 text='内置 12 家银行规则；批量模式可逐文件指定银行并合并输出',
                 font=font_ui(11), text_color=TEXT_SECONDARY, anchor='w',
                 justify='left', wraplength=520
                 ).pack(fill='x', padx=16, pady=(0, 12))

    bank_btn_row = transparent_frame(bank_card)
    bank_btn_row.pack(fill='x', padx=16, pady=(0, 16))
    ctk.CTkButton(bank_btn_row, text='单文件转换',
                  command=lambda: open_bank_converter_window(root),
                  font=font_ui(13, 'bold'), **BUTTON_PRIMARY
                  ).pack(side='left', fill='x', expand=True, padx=(0, 4))
    ctk.CTkButton(bank_btn_row, text='批量混合转换',
                  command=lambda: open_batch_converter_window(root),
                  font=font_ui(13, 'bold'), **BUTTON_SECONDARY
                  ).pack(side='left', fill='x', expand=True, padx=(4, 0))

    # 通用模板映射卡片
    generic_card = ctk.CTkFrame(root, **CARD_STYLE)
    generic_card.pack(fill='x', padx=20, pady=(0, 12))

    ctk.CTkLabel(generic_card, text='通用模板映射',
                 font=font_title(14), text_color=TEXT_PRIMARY, anchor='w'
                 ).pack(fill='x', padx=16, pady=(14, 2))
    ctk.CTkLabel(generic_card,
                 text='自定义模板 + 列映射，适用非银行流水的 Excel 转换',
                 font=font_ui(11), text_color=TEXT_SECONDARY, anchor='w'
                 ).pack(fill='x', padx=16, pady=(0, 12))

    row1 = transparent_frame(generic_card)
    row1.pack(fill='x', padx=16, pady=2)
    ctk.CTkButton(row1, text='选择模板文件', command=select_template,
                  font=font_ui(12), **BUTTON_SECONDARY
                  ).pack(side='left', fill='x', expand=True, padx=(0, 4))
    ctk.CTkButton(row1, text='设置列映射', command=set_column_mapping,
                  font=font_ui(12), **BUTTON_SECONDARY
                  ).pack(side='left', fill='x', expand=True, padx=(4, 0))

    row2 = transparent_frame(generic_card)
    row2.pack(fill='x', padx=16, pady=2)
    ctk.CTkButton(row2, text='选择生成文件路径', command=select_save_directory,
                  font=font_ui(12), **BUTTON_SECONDARY
                  ).pack(fill='x', expand=True)

    row3 = transparent_frame(generic_card)
    row3.pack(fill='x', padx=16, pady=(10, 6))
    ctk.CTkButton(row3, text='转换 Excel 文件', command=convert_excel_files,
                  font=font_ui(13, 'bold'), **BUTTON_PRIMARY
                  ).pack(fill='x', expand=True)

    row4 = transparent_frame(generic_card)
    row4.pack(fill='x', padx=16, pady=(0, 14))
    ctk.CTkButton(row4, text='历史模板', command=show_history_templates,
                  font=font_ui(12), **BUTTON_PLAIN
                  ).pack(side='left', padx=(0, 4))
    ctk.CTkButton(row4, text='历史映射', command=show_history_mappings,
                  font=font_ui(12), **BUTTON_PLAIN
                  ).pack(side='left')

    # 日志卡片
    log_card = ctk.CTkFrame(root, **CARD_STYLE)
    log_card.pack(fill='both', expand=True, padx=20, pady=(0, 20))

    log_header = transparent_frame(log_card)
    log_header.pack(fill='x', padx=16, pady=(12, 4))
    ctk.CTkLabel(log_header, text='运行日志', font=font_title(13),
                 text_color=TEXT_PRIMARY).pack(side='left')
    ctk.CTkButton(log_header, text='清空',
                  command=lambda: log_text.delete('1.0', 'end'),
                  font=font_ui(11), width=60, **BUTTON_PLAIN
                  ).pack(side='right')

    log_text = ctk.CTkTextbox(log_card, height=240, font=font_mono(11),
                              **TEXTBOX_STYLE)
    log_text.pack(fill='both', expand=True, padx=16, pady=(0, 14))

    # 日志染色 tag
    log_text.tag_config('error', foreground=RED)
    log_text.tag_config('warning', foreground=ORANGE)
    log_text.tag_config('success', foreground=GREEN)

    return root


if __name__ == '__main__':
    build_main_window().mainloop()
