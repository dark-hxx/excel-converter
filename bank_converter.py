"""银行流水转换：核心逻辑 + 单文件窗口 + 批量混合窗口（苹果风 UI）。"""
import json
import os
import re
import sys
from datetime import datetime
from decimal import Decimal
from tkinter import filedialog, ttk

import customtkinter as ctk
import pandas as pd

from apple_theme import (
    font_ui, font_title, font_mono,
    BUTTON_PRIMARY, BUTTON_SECONDARY, BUTTON_PLAIN,
    CARD_STYLE, ENTRY_STYLE, TEXTBOX_STYLE,
    BLUE, RED, GREEN, ORANGE,
    TEXT_PRIMARY, TEXT_SECONDARY, HOVER_BG,
    WINDOW_BG, CARD_BG,
    show_banner, ask_yes_no, transparent_frame,
)
from utils import (
    build_unique_save_path, center_window, make_searchable_combobox,
    open_folder, sanitize_filename_part,
)


# ---------------- 路径与常量 ----------------

if getattr(sys, 'frozen', False):
    _BASE_DIR = os.path.dirname(sys.executable)
else:
    _BASE_DIR = os.getcwd()

# 规则文件：打包后在 sys._MEIPASS，开发期在工作目录
BANK_RULES_FILE = os.path.join(
    getattr(sys, '_MEIPASS', _BASE_DIR), 'bank_rules.json'
)

# 用户上次选择（银行 + 保存目录）
LAST_CHOICE_FILE = os.path.join(_BASE_DIR, 'last_choice.json')
BANK_BUTTON_HEIGHT = 36
BANK_ACTION_BUTTON_HEIGHT = 36


def bank_button_style(style, height=BANK_BUTTON_HEIGHT):
    return dict(style, height=height)


def build_timestamped_save_path(directory, filename):
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = sanitize_filename_part(os.path.splitext(filename)[0])
    return build_unique_save_path(directory, f'{filename}_{timestamp}')


# 统一模板表头（27 列，严格匹配后端 ImportBankDetailDTO.getHeadList()）
BANK_TEMPLATE_HEADERS = [
    '交易后余额', '账户名称', '账户性质', '银行账号', '关联客户号',
    '银行流水号', '交易日期', '银行类型', '对账码', '币种',
    '摘要', '业务参考号', '借方金额', '贷方金额', '客商名称',
    '客商编号', '开户行名称', '对方账号', '对方户名', '对方开户行',
    '款项性质代码', '用途', '备注', '交易流水号', '单位名称',
    '起息日', '借/贷金额',
]

# 文件名 → 银行类型 关键字映射（按 key 长度倒序匹配，避免「中行」吞掉「中信」）
BANK_FILENAME_KEYWORDS = {
    '招商银行': '招商银行', '招商': '招商银行', 'CMB': '招商银行',
    '工商银行': '工商银行', '工行': '工商银行', 'ICBC': '工商银行',
    '农业银行': '农业银行', '农行': '农业银行', 'ABC': '农业银行',
    '中国银行': '中国银行', '中行': '中国银行', 'BOC': '中国银行',
    '交通银行': '交通银行', '交行': '交通银行', 'BCM': '交通银行',
    '中信银行': '中信银行', '中信': '中信银行', 'CITIC': '中信银行',
    '民生银行': '民生银行', '民生': '民生银行', 'CMBC': '民生银行',
    '北京银行': '北京银行', '北行': '北京银行', 'BJB': '北京银行',
    '上海银行': '上海银行', '上行': '上海银行', 'BOSC': '上海银行',
    '南京银行': '南京银行', '南京': '南京银行', 'NJB': '南京银行',
    '宁波银行': '宁波银行', '宁波': '宁波银行', 'NBB': '宁波银行',
    '农商银行': '农商银行', '农商': '农商银行', 'NSB': '农商银行',
    '江西银行': '江西银行', '江西': '江西银行', 'JXB': '江西银行',
}


# ---------------- 配置加载 ----------------

def load_bank_rules(banner_area=None):
    """加载银行规则配置。失败时通过 banner 提示并返回空字典。"""
    try:
        with open(BANK_RULES_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return {k: v for k, v in data.items() if not k.startswith('_')}
    except FileNotFoundError:
        if banner_area is not None:
            show_banner(banner_area, f'未找到银行规则文件: {BANK_RULES_FILE}', 'error', duration=0)
        return {}
    except Exception as e:
        if banner_area is not None:
            show_banner(banner_area, f'加载银行规则失败: {e}', 'error', duration=0)
        return {}


def load_last_choice():
    try:
        with open(LAST_CHOICE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def save_last_choice(bank, save_dir_path):
    try:
        with open(LAST_CHOICE_FILE, 'w', encoding='utf-8') as f:
            json.dump({'bank': bank, 'save_dir': save_dir_path},
                      f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def guess_bank_by_filename(filename, available_banks):
    """根据文件名启发式猜银行；猜不到返回 None。"""
    name = os.path.basename(filename)
    for kw in sorted(BANK_FILENAME_KEYWORDS.keys(), key=len, reverse=True):
        if kw in name:
            bank = BANK_FILENAME_KEYWORDS[kw]
            if bank in available_banks:
                return bank
    return None


# ---------------- 转换核心（业务逻辑，UI 无关） ----------------

def _bank_log(log_widget, msg):
    """日志输出，按内容自动染色。"""
    if log_widget is None:
        return
    level = None
    if '失败' in msg or '错误' in msg or '异常' in msg:
        level = 'error'
    elif '跳过' in msg or '警告' in msg:
        level = 'warning'
    elif '成功' in msg or '完成' in msg:
        level = 'success'

    line = msg + '\n'
    try:
        if level:
            log_widget.insert('end', line, level)
        else:
            log_widget.insert('end', line)
        log_widget.see('end')
        log_widget.update_idletasks()
    except Exception:
        # 兼容传入非 ctk 控件（保底无染色）
        log_widget.insert('end', line)
        log_widget.see('end')


def parse_amount(val):
    """解析金额：去千分位、去货币符号；空/横线/异常返回 None"""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s in ('-', '--', 'nan', 'None'):
        return None
    s = s.replace(',', '').replace(' ', '').replace('¥', '').replace('￥', '')
    try:
        return Decimal(s)
    except Exception:
        return None


def _amount_to_str(d):
    if d is None:
        return ''
    s = format(d, 'f')
    if '.' in s:
        s = s.rstrip('0').rstrip('.')
    return s or '0'


def extract_account_info(file_path, rule):
    """从顶部行/单元格提取账户信息。"""
    extract_cfg = rule.get('account_extract')
    if not extract_cfg:
        return {}

    from openpyxl import load_workbook
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        result = {}
        for tpl_field, spec in extract_cfg.items():
            cell_addr = spec.get('cell')
            if not cell_addr:
                continue
            raw = ws[cell_addr].value
            if raw is None:
                continue
            text = str(raw)
            if spec.get('strip', False):
                text = text.strip().strip('\t').strip()
            if spec.get('mode', 'cell') == 'regex':
                m = re.search(spec.get('pattern', ''), text)
                if m and m.groups():
                    text = m.group(1).strip()
            result[tpl_field] = text
        wb.close()
        return result
    except Exception:
        return {}


def _get_cell(row, col_name):
    if col_name is None:
        return None
    try:
        if (col_name in row.index) if hasattr(row, 'index') else (col_name in row):
            v = row[col_name]
            if pd.isna(v):
                return None
            return v
    except Exception:
        pass
    return None


def build_date_value(row, spec):
    """根据 spec 构建日期字符串。spec 可以是 str (列名) 或 dict。"""
    if isinstance(spec, str):
        val = _get_cell(row, spec)
        return '' if val is None else str(val).strip()

    source = spec.get('source')
    in_fmt = spec.get('in_fmt')
    out_fmt = spec.get('out_fmt')
    join_str = spec.get('join', ' ')

    if isinstance(source, list):
        parts = []
        for col in source:
            v = _get_cell(row, col)
            if v is None:
                parts.append('')
                continue
            s = str(v).strip()
            if spec.get('date_only_first') and len(parts) == 0 and ' ' in s:
                s = s.split(' ')[0]
            parts.append(s)
        raw = join_str.join([p for p in parts if p])
    else:
        v = _get_cell(row, source)
        raw = '' if v is None else str(v).strip()

    if not raw:
        return ''

    if spec.get('pad_seconds') and raw.count(':') == 1:
        raw = raw + ':00'

    if not in_fmt or not out_fmt:
        return raw

    try:
        return datetime.strptime(raw, in_fmt).strftime(out_fmt)
    except Exception:
        return raw


def resolve_field(row, spec):
    if spec is None:
        return ''
    if isinstance(spec, str):
        v = _get_cell(row, spec)
        return '' if v is None else str(v).strip()
    src = spec.get('source')
    v = _get_cell(row, src)
    if v is None:
        return ''
    s = str(v).strip()
    if spec.get('strip_spaces'):
        s = s.replace(' ', '')
    return s


def normalize_debit_credit(row, rule):
    """借贷归一，返回 (借方金额字符串, 贷方金额字符串)"""
    mode = rule.get('debit_credit_mode', 'two_columns')
    cm = rule.get('column_mapping', {})

    if mode == 'two_columns':
        jie_spec = cm.get('借方金额')
        dai_spec = cm.get('贷方金额')
        jie_col = jie_spec if isinstance(jie_spec, str) else (jie_spec or {}).get('source')
        dai_col = dai_spec if isinstance(dai_spec, str) else (dai_spec or {}).get('source')
        jie = parse_amount(_get_cell(row, jie_col))
        dai = parse_amount(_get_cell(row, dai_col))
        return _amount_to_str(jie), _amount_to_str(dai)

    if mode == 'signed_amount':
        amt = parse_amount(_get_cell(row, rule.get('amount_column')))
        if amt is None or amt == 0:
            return '', ''
        if amt < 0:
            return _amount_to_str(-amt), ''
        return '', _amount_to_str(amt)

    if mode == 'marker_column':
        mc = rule.get('marker_column', {})
        marker = _get_cell(row, mc.get('col'))
        marker_str = '' if marker is None else str(marker).strip()
        amt = parse_amount(_get_cell(row, mc.get('amount_col')))
        if amt is None:
            return '', ''
        if marker_str == mc.get('jie_value'):
            return _amount_to_str(amt), ''
        if marker_str == mc.get('dai_value'):
            return '', _amount_to_str(amt)
        return '', ''

    return '', ''


def _transform_one_row(row, rule, account_info, bank_type_code):
    """把源文件一行转成目标模板的 dict（27 列）。失败字段留空。"""
    cm = rule.get('column_mapping', {})
    fixed = rule.get('fixed_values', {})

    out = {h: '' for h in BANK_TEMPLATE_HEADERS}

    for tpl_field, spec in cm.items():
        if tpl_field in ('交易日期', '起息日'):
            out[tpl_field] = build_date_value(row, spec)
        elif tpl_field in ('借方金额', '贷方金额'):
            pass
        elif tpl_field == '交易后余额':
            bal_col = spec if isinstance(spec, str) else spec.get('source')
            out[tpl_field] = _amount_to_str(parse_amount(_get_cell(row, bal_col)))
        else:
            out[tpl_field] = resolve_field(row, spec)

    jie, dai = normalize_debit_credit(row, rule)
    out['借方金额'] = jie
    out['贷方金额'] = dai

    for k, v in account_info.items():
        if not out.get(k):
            out[k] = v
    for k, v in fixed.items():
        if not out.get(k):
            out[k] = str(v)
    if not out.get('银行类型'):
        out['银行类型'] = bank_type_code

    return out


def convert_bank_rows(file_path, rule, log_widget=None):
    """读取并转换一个文件，返回 (rows_list, skipped_count, error_msg)。"""
    try:
        header_row = rule.get('header_row', 1)
        df = pd.read_excel(file_path, sheet_name=0,
                           header=header_row - 1, dtype=str)
    except Exception as e:
        return [], 0, f'读取失败: {e}'

    df = df.loc[:, df.columns.notna()]
    df.columns = [str(c).strip() for c in df.columns]
    _bank_log(log_widget, f'  共 {len(df)} 行待处理')

    account_info = extract_account_info(file_path, rule)
    if account_info:
        _bank_log(log_widget, f'  顶部账户信息: {account_info}')

    bank_type_code = rule.get('bank_type_code', '')
    rows_out = []
    skipped = 0

    for idx, row in df.iterrows():
        out = _transform_one_row(row, rule, account_info, bank_type_code)
        if not out['交易后余额'] or not out['交易日期']:
            skipped += 1
            _bank_log(log_widget,
                      f'  跳过第 {idx + header_row + 1} 行：余额或交易日期为空')
            continue
        rows_out.append(out)

    return rows_out, skipped, ''


def _write_bank_xlsx(rows, save_path):
    """把 rows 写入指定 xlsx，所有单元格按文本格式。"""
    from openpyxl import load_workbook
    from openpyxl.styles import numbers

    out_df = pd.DataFrame(rows, columns=BANK_TEMPLATE_HEADERS)
    out_df.to_excel(save_path, index=False, sheet_name='Sheet1', engine='openpyxl')
    wb = load_workbook(save_path)
    ws = wb.active
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in r:
            cell.number_format = numbers.FORMAT_TEXT
    wb.save(save_path)
    wb.close()


def convert_bank_file(file_path, bank_name, rule, save_dir_path, log_widget):
    """单文件转换。成功返回 True。"""
    _bank_log(log_widget, f'开始处理 [{bank_name}] {file_path}')

    rows, skipped, err = convert_bank_rows(file_path, rule, log_widget)
    if err:
        _bank_log(log_widget, f'  {err}')
        return False
    if not rows:
        _bank_log(log_widget, '  无有效数据行，未生成文件')
        return False

    file_name = os.path.splitext(os.path.basename(file_path))[0]
    save_path = build_timestamped_save_path(
        save_dir_path, f'{file_name}_{bank_name}_统一格式'
    )

    try:
        _write_bank_xlsx(rows, save_path)
    except PermissionError:
        _bank_log(log_widget, f'  保存失败：{save_path} 被占用，请关闭后重试')
        return False
    except Exception as e:
        _bank_log(log_widget, f'  保存失败: {e}')
        return False

    _bank_log(log_widget,
              f'  完成：输出 {len(rows)} 行，跳过 {skipped} 行 → {save_path}')
    return True


# ---------------- 预览窗口 ----------------

def preview_bank_file(file_path, bank_name, rule, win_parent, log_widget):
    """读取前 5 行做转换预演，弹窗用 Treeview 展示关键字段。"""
    try:
        header_row = rule.get('header_row', 1)
        df = pd.read_excel(file_path, sheet_name=0,
                           header=header_row - 1, dtype=str, nrows=5)
    except Exception as e:
        _bank_log(log_widget, f'预览失败: {e}')
        return

    df = df.loc[:, df.columns.notna()]
    df.columns = [str(c).strip() for c in df.columns]

    account_info = extract_account_info(file_path, rule)
    bank_type_code = rule.get('bank_type_code', '')

    preview_rows = [
        _transform_one_row(row, rule, account_info, bank_type_code)
        for _, row in df.iterrows()
    ]
    if not preview_rows:
        _bank_log(log_widget, '无数据可预览')
        return

    preview_cols = ['交易日期', '银行账号', '银行类型', '币种', '借方金额',
                    '贷方金额', '交易后余额', '对方账号', '对方户名', '摘要']
    col_widths = {
        '交易日期': 150, '银行账号': 180, '银行类型': 70, '币种': 50,
        '借方金额': 110, '贷方金额': 110, '交易后余额': 130,
        '对方账号': 180, '对方户名': 200, '摘要': 180,
    }

    pw = ctk.CTkToplevel(win_parent)
    pw.title(f'预览：{bank_name} - 前 {len(preview_rows)} 行')
    pw.configure(fg_color=WINDOW_BG)
    center_window(pw, 1320, 400)
    pw.transient(win_parent)
    pw.grab_set()

    ctk.CTkLabel(pw, text=f'{bank_name} · 前 {len(preview_rows)} 行预览',
                 font=font_title(14), text_color=TEXT_PRIMARY, anchor='w'
                 ).pack(fill='x', padx=20, pady=(14, 8))

    table_card = ctk.CTkFrame(pw, **CARD_STYLE)
    table_card.pack(fill='both', expand=True, padx=16, pady=(0, 12))

    tree = ttk.Treeview(table_card, columns=preview_cols, show='headings',
                        height=len(preview_rows))
    for col in preview_cols:
        tree.heading(col, text=col)
        tree.column(col, width=col_widths.get(col, 120), anchor='w', stretch=False)
    for r in preview_rows:
        tree.insert('', 'end', values=[r.get(c, '') for c in preview_cols])

    vsb = ttk.Scrollbar(table_card, orient='vertical', command=tree.yview)
    hsb = ttk.Scrollbar(table_card, orient='horizontal', command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

    tree.grid(row=0, column=0, sticky='nsew', padx=(8, 0), pady=(8, 0))
    vsb.grid(row=0, column=1, sticky='ns', pady=(8, 0))
    hsb.grid(row=1, column=0, sticky='ew', padx=(8, 0))
    table_card.rowconfigure(0, weight=1)
    table_card.columnconfigure(0, weight=1)

    ctk.CTkButton(pw, text='关闭', command=pw.destroy,
                  width=120, font=font_ui(12), **bank_button_style(BUTTON_PLAIN)
                  ).pack(pady=(0, 14))


# ---------------- 单文件转换窗口 ----------------

def open_bank_converter_window(parent_root):
    win = ctk.CTkToplevel(parent_root)
    win.title('银行流水转换')
    win.configure(fg_color=WINDOW_BG)
    center_window(win, 660, 720)
    win.transient(parent_root)
    win.grab_set()
    win.focus_set()

    # 顶部 banner 区
    banner_area = transparent_frame(win)
    banner_area.pack(fill='x', side='top')

    rules = load_bank_rules(banner_area)
    if not rules:
        return

    last = load_last_choice()
    bank_names = list(rules.keys())
    last_bank = last.get('bank') if last.get('bank') in bank_names else bank_names[0]
    last_dir = last.get('save_dir') or ''

    # 标题
    ctk.CTkLabel(win, text='银行流水转换', font=font_title(16),
                 text_color=TEXT_PRIMARY, anchor='w'
                 ).pack(fill='x', padx=20, pady=(14, 8))

    # 表单卡片
    form_card = ctk.CTkFrame(win, **CARD_STYLE)
    form_card.pack(fill='x', padx=16, pady=(0, 10))

    bank_var = ctk.StringVar()
    file_var = ctk.StringVar()
    dir_var = ctk.StringVar(value=last_dir)

    def _row(parent, label_text, row_idx):
        ctk.CTkLabel(parent, text=label_text, font=font_ui(12, 'bold'),
                     text_color=TEXT_SECONDARY, width=80, anchor='w'
                     ).grid(row=row_idx, column=0, sticky='w', padx=(16, 8), pady=10)

    _row(form_card, '银行类型', 0)
    bank_combo, _ = make_searchable_combobox(form_card, bank_names,
                                             textvariable=bank_var, width=24)
    bank_combo.grid(row=0, column=1, sticky='w', padx=(0, 8), pady=10)
    bank_combo.set(last_bank)

    _row(form_card, '源文件', 1)
    file_label = ctk.CTkLabel(form_card, textvariable=file_var,
                              text_color=TEXT_PRIMARY, font=font_ui(11),
                              anchor='w', justify='left', wraplength=380)
    file_label.grid(row=1, column=1, sticky='w', padx=(0, 8), pady=10)

    def pick_file():
        p = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx;*.xls')])
        if p:
            file_var.set(p)

    ctk.CTkButton(form_card, text='选择文件', command=pick_file,
                  width=88, font=font_ui(11), **bank_button_style(BUTTON_PLAIN)
                  ).grid(row=1, column=2, padx=(0, 16), pady=10)

    _row(form_card, '保存目录', 2)
    ctk.CTkLabel(form_card, textvariable=dir_var, text_color=TEXT_PRIMARY,
                 font=font_ui(11), anchor='w', justify='left', wraplength=380
                 ).grid(row=2, column=1, sticky='w', padx=(0, 8), pady=(10, 16))

    def pick_dir():
        d = filedialog.askdirectory()
        if d:
            dir_var.set(d)

    ctk.CTkButton(form_card, text='选择目录', command=pick_dir,
                  width=88, font=font_ui(11), **bank_button_style(BUTTON_PLAIN)
                  ).grid(row=2, column=2, padx=(0, 16), pady=(10, 16))

    form_card.columnconfigure(1, weight=1)

    # 操作按钮区
    btn_row = transparent_frame(win)
    btn_row.pack(fill='x', padx=16, pady=(0, 10))
    buttons = {}

    def do_preview():
        if not bank_var.get():
            show_banner(banner_area, '请选择银行类型', 'warning')
            return
        if not file_var.get():
            show_banner(banner_area, '请选择源文件', 'warning')
            return
        rule = rules.get(bank_var.get())
        if rule:
            preview_bank_file(file_var.get(), bank_var.get(), rule, win, log_widget)

    def do_convert():
        bank_name = bank_var.get()
        file_path = file_var.get()
        save_path = dir_var.get()

        if not bank_name:
            show_banner(banner_area, '请选择银行类型', 'warning')
            return
        if not file_path:
            show_banner(banner_area, '请选择源文件', 'warning')
            return
        if not save_path:
            show_banner(banner_area, '请选择保存目录', 'warning')
            return

        rule = rules.get(bank_name)
        if not rule:
            show_banner(banner_area, f'未找到银行规则：{bank_name}', 'error')
            return

        log_widget.delete('1.0', 'end')
        buttons['preview'].configure(state='disabled')
        buttons['convert'].configure(state='disabled', text='处理中...')
        win.update_idletasks()
        try:
            ok = convert_bank_file(file_path, bank_name, rule, save_path, log_widget)
        finally:
            buttons['preview'].configure(state='normal')
            buttons['convert'].configure(state='normal', text='开始转换')

        if ok:
            save_last_choice(bank_name, save_path)
            if ask_yes_no(win, '完成', '转换完成，是否打开输出目录？',
                          yes_text='打开目录', no_text='关闭'):
                open_folder(save_path)
        else:
            show_banner(banner_area, '转换未完成，请查看日志', 'warning')

    buttons['preview'] = ctk.CTkButton(
        btn_row, text='预览前 5 行', command=do_preview,
        font=font_ui(12), **bank_button_style(BUTTON_SECONDARY))
    buttons['preview'].pack(side='left', fill='x', expand=True, padx=(0, 4))

    buttons['convert'] = ctk.CTkButton(
        btn_row, text='开始转换', command=do_convert,
        font=font_ui(13, 'bold'), **bank_button_style(BUTTON_PRIMARY))
    buttons['convert'].pack(side='left', fill='x', expand=True, padx=(4, 0))

    # 日志卡片
    log_card = ctk.CTkFrame(win, **CARD_STYLE)
    log_card.pack(fill='both', expand=True, padx=16, pady=(0, 16))

    log_header = transparent_frame(log_card)
    log_header.pack(fill='x', padx=14, pady=(10, 4))
    ctk.CTkLabel(log_header, text='日志', font=font_title(12),
                 text_color=TEXT_PRIMARY).pack(side='left')
    ctk.CTkButton(log_header, text='清空',
                  command=lambda: log_widget.delete('1.0', 'end'),
                  font=font_ui(11), width=60, **bank_button_style(BUTTON_PLAIN)
                  ).pack(side='right')

    log_widget = ctk.CTkTextbox(log_card, height=240, font=font_mono(11),
                                **TEXTBOX_STYLE)
    log_widget.pack(fill='both', expand=True, padx=14, pady=(0, 12))

    log_widget.tag_config('error', foreground=RED)
    log_widget.tag_config('warning', foreground=ORANGE)
    log_widget.tag_config('success', foreground=GREEN)


# ---------------- 批量混合窗口 ----------------

def open_batch_converter_window(parent_root):
    win = ctk.CTkToplevel(parent_root)
    win.title('批量混合转换')
    win.configure(fg_color=WINDOW_BG)
    center_window(win, 960, 860)
    win.transient(parent_root)
    win.grab_set()
    win.focus_set()

    banner_area = transparent_frame(win)
    banner_area.pack(fill='x', side='top')

    rules = load_bank_rules(banner_area)
    if not rules:
        return

    bank_names = list(rules.keys())
    last = load_last_choice()
    last_dir = last.get('save_dir') or ''

    path_by_iid = {}

    # 标题
    ctk.CTkLabel(win, text='批量混合转换', font=font_title(16),
                 text_color=TEXT_PRIMARY, anchor='w'
                 ).pack(fill='x', padx=20, pady=(14, 4))
    ctk.CTkLabel(win, text='可一次添加多家银行的多个文件，逐文件指定银行类型，合并到一个输出文件',
                 font=font_ui(11), text_color=TEXT_SECONDARY, anchor='w'
                 ).pack(fill='x', padx=20, pady=(0, 8))

    # 底部按钮区：提前 pack 占位 + 锁定高度，确保即使内容溢出也始终可见
    action_row = transparent_frame(win, height=68)
    action_row.pack(side='bottom', fill='x', padx=16, pady=(0, 14))
    action_row.pack_propagate(False)

    # 文件列表卡片
    files_card = ctk.CTkFrame(win, **CARD_STYLE)
    files_card.pack(fill='both', expand=True, padx=16, pady=(0, 10))

    file_btn_row = transparent_frame(files_card)
    file_btn_row.pack(fill='x', padx=12, pady=(12, 4))

    list_container = transparent_frame(files_card)
    list_container.pack(fill='both', expand=True, padx=12, pady=(4, 4))

    list_cols = ('#', '文件名', '银行类型', '状态')
    col_widths = {'#': 40, '文件名': 380, '银行类型': 140, '状态': 160}

    tree = ttk.Treeview(list_container, columns=list_cols, show='headings', height=10)
    for c in list_cols:
        tree.heading(c, text=c)
        tree.column(c, width=col_widths.get(c, 100),
                    anchor='center' if c in ('#', '状态') else 'w',
                    stretch=(c == '文件名'))

    vsb = ttk.Scrollbar(list_container, orient='vertical', command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.grid(row=0, column=0, sticky='nsew')
    vsb.grid(row=0, column=1, sticky='ns')
    list_container.rowconfigure(0, weight=1)
    list_container.columnconfigure(0, weight=1)

    def _renumber():
        for i, iid in enumerate(tree.get_children(), 1):
            vals = list(tree.item(iid, 'values'))
            vals[0] = i
            tree.item(iid, values=vals)

    def add_files():
        paths = filedialog.askopenfilenames(filetypes=[('Excel files', '*.xlsx;*.xls')])
        for p in paths:
            bank = guess_bank_by_filename(p, bank_names) or bank_names[0]
            iid = tree.insert('', 'end', values=(
                len(path_by_iid) + 1, os.path.basename(p), bank, '待转换'))
            path_by_iid[iid] = p
        _renumber()

    def remove_selected():
        for iid in tree.selection():
            path_by_iid.pop(iid, None)
            tree.delete(iid)
        _renumber()

    def clear_all():
        for iid in list(path_by_iid):
            tree.delete(iid)
        path_by_iid.clear()

    ctk.CTkButton(file_btn_row, text='+ 添加文件', command=add_files,
                  font=font_ui(12), width=110, **bank_button_style(BUTTON_PRIMARY)
                  ).pack(side='left', padx=(0, 4))
    ctk.CTkButton(file_btn_row, text='- 移除选中', command=remove_selected,
                  font=font_ui(12), width=100, **bank_button_style(BUTTON_SECONDARY)
                  ).pack(side='left', padx=4)
    ctk.CTkButton(file_btn_row, text='清空', command=clear_all,
                  font=font_ui(12), width=80, **bank_button_style(BUTTON_PLAIN)
                  ).pack(side='left', padx=4)
    ctk.CTkLabel(file_btn_row, text='Ctrl/Shift 多选',
                 font=font_ui(11), text_color=TEXT_SECONDARY
                 ).pack(side='left', padx=12)

    # 修改选中行银行
    edit_row = transparent_frame(files_card)
    edit_row.pack(fill='x', padx=12, pady=(4, 12))
    ctk.CTkLabel(edit_row, text='选中行改为：', font=font_ui(12),
                 text_color=TEXT_SECONDARY).pack(side='left')
    edit_var = ctk.StringVar(value=bank_names[0])
    edit_combo, _ = make_searchable_combobox(edit_row, bank_names,
                                             textvariable=edit_var, width=18)
    edit_combo.pack(side='left', padx=6)

    def apply_bank():
        sels = tree.selection()
        if not sels:
            show_banner(banner_area, '请先在列表中选中行（可多选）', 'info')
            return
        bank = edit_var.get()
        for iid in sels:
            vals = list(tree.item(iid, 'values'))
            vals[2] = bank
            vals[3] = '待转换'
            tree.item(iid, values=vals)

    ctk.CTkButton(edit_row, text='应用', command=apply_bank,
                  font=font_ui(12), width=70, **bank_button_style(BUTTON_SECONDARY)
                  ).pack(side='left', padx=4)

    # 输出设置卡片
    out_card = ctk.CTkFrame(win, **CARD_STYLE)
    out_card.pack(fill='x', padx=16, pady=(0, 10))

    ctk.CTkLabel(out_card, text='输出目录', font=font_ui(12, 'bold'),
                 text_color=TEXT_SECONDARY, width=80, anchor='w'
                 ).grid(row=0, column=0, sticky='w', padx=(16, 8), pady=(14, 6))

    out_dir_var = ctk.StringVar(value=last_dir)
    ctk.CTkLabel(out_card, textvariable=out_dir_var, text_color=TEXT_PRIMARY,
                 font=font_ui(11), anchor='w', justify='left', wraplength=620
                 ).grid(row=0, column=1, sticky='w', padx=(0, 8), pady=(14, 6))

    def pick_out_dir():
        d = filedialog.askdirectory()
        if d:
            out_dir_var.set(d)

    ctk.CTkButton(out_card, text='浏览', command=pick_out_dir,
                  width=70, font=font_ui(11), **bank_button_style(BUTTON_PLAIN)
                  ).grid(row=0, column=2, padx=(0, 16), pady=(14, 6))

    ctk.CTkLabel(out_card, text='文件名', font=font_ui(12, 'bold'),
                 text_color=TEXT_SECONDARY, width=80, anchor='w'
                 ).grid(row=1, column=0, sticky='w', padx=(16, 8), pady=(6, 14))

    out_name_var = ctk.StringVar(value='合并流水.xlsx')
    ctk.CTkEntry(out_card, textvariable=out_name_var, width=440, **ENTRY_STYLE
                 ).grid(row=1, column=1, sticky='w', padx=(0, 8), pady=(6, 14))

    out_card.columnconfigure(1, weight=1)

    # 日志卡片
    log_card = ctk.CTkFrame(win, **CARD_STYLE)
    log_card.pack(fill='both', expand=True, padx=16, pady=(0, 10))

    log_header = transparent_frame(log_card)
    log_header.pack(fill='x', padx=14, pady=(10, 4))
    ctk.CTkLabel(log_header, text='日志', font=font_title(12),
                 text_color=TEXT_PRIMARY).pack(side='left')
    ctk.CTkButton(log_header, text='清空',
                  command=lambda: log_widget.delete('1.0', 'end'),
                  font=font_ui(11), width=60, **bank_button_style(BUTTON_PLAIN)
                  ).pack(side='right')

    log_widget = ctk.CTkTextbox(log_card, height=240, font=font_mono(11),
                                **TEXTBOX_STYLE)
    log_widget.pack(fill='both', expand=True, padx=14, pady=(0, 12))

    log_widget.tag_config('error', foreground=RED)
    log_widget.tag_config('warning', foreground=ORANGE)
    log_widget.tag_config('success', foreground=GREEN)

    def do_batch_convert():
        if not path_by_iid:
            show_banner(banner_area, '请先添加文件', 'warning')
            return
        save_dir_path = out_dir_var.get().strip()
        if not save_dir_path:
            show_banner(banner_area, '请选择输出目录', 'warning')
            return
        out_name = out_name_var.get().strip()
        if not out_name:
            show_banner(banner_area, '请填写输出文件名', 'warning')
            return
        if not out_name.lower().endswith('.xlsx'):
            out_name += '.xlsx'

        log_widget.delete('1.0', 'end')
        convert_btn.configure(state='disabled', text='处理中...')
        win.update_idletasks()

        merged = []
        try:
            for iid in tree.get_children():
                vals = list(tree.item(iid, 'values'))
                idx_no, fname, bank, _ = vals
                fpath = path_by_iid[iid]

                tree.item(iid, values=(idx_no, fname, bank, '处理中...'))
                win.update_idletasks()
                _bank_log(log_widget, f'\n[{idx_no}] [{bank}] {fname}')

                rule = rules.get(bank)
                if not rule:
                    tree.item(iid, values=(idx_no, fname, bank, '失败'))
                    _bank_log(log_widget, f'  未找到规则：{bank}，整批终止')
                    show_banner(banner_area, f'未找到银行规则：{bank}，整批终止', 'error')
                    return

                rows, skipped, err = convert_bank_rows(fpath, rule, log_widget)
                if err:
                    tree.item(iid, values=(idx_no, fname, bank, '失败'))
                    _bank_log(log_widget, f'  {err}，整批终止')
                    show_banner(banner_area,
                                f'[{fname}] 转换失败：{err}（已合并 {len(merged)} 行未写入）',
                                'error')
                    return
                if not rows:
                    tree.item(iid, values=(idx_no, fname, bank, '无有效行'))
                    _bank_log(log_widget, '  无有效行，跳过但继续后续文件')
                    continue

                merged.extend(rows)
                tree.item(iid, values=(idx_no, fname, bank, f'成功 ({len(rows)} 行)'))
                _bank_log(log_widget, f'  汇入 {len(rows)} 行（跳过 {skipped}）')

            if not merged:
                _bank_log(log_widget, '\n所有文件均无有效数据，未生成合并文件')
                show_banner(banner_area, '所有文件均无有效数据，未生成合并文件', 'warning')
                return

            output_name = os.path.splitext(out_name)[0]
            save_path = build_timestamped_save_path(save_dir_path, output_name)
            try:
                _write_bank_xlsx(merged, save_path)
            except PermissionError:
                msg = f'{save_path} 被占用，请关闭后重试'
                _bank_log(log_widget, f'\n保存失败：{msg}')
                show_banner(banner_area, f'保存失败：{msg}', 'error')
                return
            except Exception as e:
                _bank_log(log_widget, f'\n保存失败：{e}')
                show_banner(banner_area, f'保存失败：{e}', 'error')
                return

            _bank_log(log_widget,
                      f'\n========== 全部完成：合并 {len(merged)} 行 → {save_path} ==========')
            save_last_choice(bank_names[0], save_dir_path)
            if ask_yes_no(win, '完成',
                          f'共合并 {len(merged)} 行到 {os.path.basename(save_path)}\n是否打开输出目录？',
                          yes_text='打开目录', no_text='关闭'):
                open_folder(save_dir_path)
        finally:
            convert_btn.configure(state='normal', text='开始转换')

    convert_btn = ctk.CTkButton(
        action_row, text='开始转换', command=do_batch_convert,
        font=font_ui(14, 'bold'), **bank_button_style(BUTTON_PRIMARY))
    convert_btn.pack(fill='x')
